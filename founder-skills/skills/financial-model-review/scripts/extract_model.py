#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///
"""Extract structured data from Excel (.xlsx) or CSV files.

Usage:
    python extract_model.py --file model.xlsx --pretty
    python extract_model.py --file data.csv -o model_data.json
    echo '{"sheets": [...]}' | python extract_model.py --stdin

Output: JSON with structure:
    {
      "sheets": [
        {
          "name": str,
          "headers": [str],
          "rows": [[value]],
          "detected_type": str|null,
          "periodicity": str|null,
          "row_count": int,
          "col_count": int,
          "pre_header_rows": [[value]],
          "cell_refs": {str: str}
        }
      ],
      "source_format": str,
      "source_file": str,
      "periodicity_summary": {str: int}
    }
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import warnings
from typing import Any


def _write_output(data: str, output_path: str | None, *, summary: dict[str, Any] | None = None) -> None:
    if output_path:
        abs_path = os.path.abspath(output_path)
        parent = os.path.dirname(abs_path)
        if parent == "/":
            print(f"Error: output path resolves to root directory: {abs_path}", file=sys.stderr)
            sys.exit(1)
        os.makedirs(parent, exist_ok=True)
        with open(abs_path, "w", encoding="utf-8") as f:
            f.write(data)
        receipt: dict[str, Any] = {"ok": True, "path": abs_path, "bytes": len(data.encode("utf-8"))}
        if summary:
            receipt.update(summary)
        sys.stdout.write(json.dumps(receipt, separators=(",", ":")) + "\n")
    else:
        sys.stdout.write(data)


# Tab name heuristics for detecting sheet purpose
_TAB_PATTERNS: dict[str, list[str]] = {
    "assumptions": ["assumption", "input", "driver", "parameter"],
    "revenue": ["revenue", "sales", "arr", "mrr", "income"],
    "expenses": ["expense", "opex", "cost", "headcount", "hiring", "payroll"],
    "cash": ["cash", "runway", "burn", "balance"],
    "pnl": ["p&l", "pnl", "profit", "loss", "income statement"],
    "summary": ["summary", "dashboard", "overview", "kpi"],
    "scenarios": ["scenario", "sensitivity", "case"],
}


def _detect_tab_type(name: str) -> str | None:
    lower = name.lower().strip()
    for tab_type, patterns in _TAB_PATTERNS.items():
        for pat in patterns:
            if pat in lower:
                return tab_type
    return None


def _safe_value(val: Any) -> Any:
    """Convert cell value to JSON-serializable type."""
    if val is None:
        return None
    if isinstance(val, (int, float, bool)):
        return val
    return str(val)


# ---------------------------------------------------------------------------
# Periodicity detection
# ---------------------------------------------------------------------------

# Month-range patterns MUST be checked before single-month to avoid
# misclassifying "Jan-Mar" as monthly.
_MONTH_NAMES = (
    r"jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?"
    r"|jul(?:y)?|aug(?:ust)?|sep(?:tember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?"
)

_QUARTERLY_PATTERNS: list[re.Pattern[str]] = [
    # Q1 2024, Q1-24, Q1 - 24, Q1'24
    re.compile(r"\bQ[1-4]\b", re.IGNORECASE),
    # 1Q24, 1Q2024
    re.compile(r"\b[1-4]Q\d{2,4}\b", re.IGNORECASE),
    # Jan-Mar 2024, January-March, Jan-Mar
    re.compile(
        rf"\b({_MONTH_NAMES})\s*[-–]\s*({_MONTH_NAMES})\b",
        re.IGNORECASE,
    ),
]

_ANNUAL_PATTERNS: list[re.Pattern[str]] = [
    # FY2024, FY24, FY 2024
    re.compile(r"\bFY\s*\d{2,4}\b", re.IGNORECASE),
    # H1 2024, H2, 1H24, 2H24
    re.compile(r"\b[12]H\d{2,4}\b|\bH[12]\b", re.IGNORECASE),
]

_MONTHLY_PATTERNS: list[re.Pattern[str]] = [
    # Jan 2024, January 24, Jan-24, Jan '24
    re.compile(
        rf"\b({_MONTH_NAMES})\s*[-–'/]?\s*\d{{2,4}}\b",
        re.IGNORECASE,
    ),
    # 2024-01, 2024-01-01
    re.compile(r"\b20\d{2}-(?:0[1-9]|1[0-2])\b"),
]


def _classify_header(header: str) -> str | None:
    """Classify a single column header as monthly/quarterly/annual or None."""
    # Check quarterly first (month-range before single-month)
    for pat in _QUARTERLY_PATTERNS:
        if pat.search(header):
            return "quarterly"
    for pat in _ANNUAL_PATTERNS:
        if pat.search(header):
            return "annual"
    for pat in _MONTHLY_PATTERNS:
        if pat.search(header):
            return "monthly"
    return None


def detect_periodicity(headers: list[str]) -> str:
    """Detect periodicity from column headers via majority vote.

    Skips the first column (typically row labels). Returns one of:
    monthly, quarterly, annual, unknown.
    """
    classifications: list[str] = []
    for h in headers[1:]:  # skip first column (row labels)
        c = _classify_header(h)
        if c is not None:
            classifications.append(c)

    if not classifications:
        return "unknown"

    # Majority vote
    from collections import Counter

    counts = Counter(classifications)
    winner, _ = counts.most_common(1)[0]
    return winner


# Spreadsheet error tokens. openpyxl with data_only=True returns the cached
# computed value, so a broken cell arrives as this literal string — the signal is
# already in the extracted values, it just needs counting. STRUCT_08 is scored on
# exactly these, and without a summary the assessor (which reads inputs.json, not
# this file) has no way to see them.
_SPREADSHEET_ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!")


def _count_structural_errors(sheets: list[dict[str, Any]]) -> dict[str, int]:
    """Tally spreadsheet error cells across every extracted sheet."""
    counts: dict[str, int] = {}
    for sheet in sheets:
        for row in sheet.get("rows", []) or []:
            for cell in row:
                if isinstance(cell, str):
                    token = cell.strip().upper()
                    if token in _SPREADSHEET_ERROR_TOKENS:
                        counts[token] = counts.get(token, 0) + 1
    return counts


def _periodicity_summary(sheets: list[dict[str, Any]]) -> str:
    """Compute top-level periodicity summary from per-sheet values."""
    values: set[str] = {s["periodicity"] for s in sheets if s.get("periodicity") != "unknown"}
    if not values:
        return "unknown"
    if len(values) == 1:
        return next(iter(values))
    return "mixed"


_MAX_HEADER_SCAN = 10  # Scan first N rows for header detection

# Patterns that identify a cell as a DATA value rather than a header label.
# A row that is predominantly data (numeric or date-like) is not a banner/header;
# banner rows in real exports are text-dominant ("Acme Summary…", "Assumptions").
_ISO_MONTH_RE = re.compile(r"^\d{4}-(0[1-9]|1[0-2])$")  # 2025-01 style
_ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")  # 2025-01-01 full date


def _cell_is_data(val: Any) -> bool:
    """Return True when a cell looks like a data value, not a column label.

    Numeric values (int/float, not bool) and ISO date/month strings are data.
    Boolean openpyxl cells are not data here (they're usually labels or flags).
    """
    if isinstance(val, bool):
        return False
    if isinstance(val, (int, float)):
        return True
    if isinstance(val, str):
        s = val.strip()
        if _ISO_MONTH_RE.match(s) or _ISO_DATE_RE.match(s):
            return True
    return False


def _row_is_predominantly_data(row: list[Any]) -> bool:
    """Return True when the majority of non-null cells in a row are data values.

    A row where most cells are numeric or date-like is a data row, not a header.
    Used as a disqualification guard in _find_header_row so that leading data rows
    with period-matching content (e.g. '2025-01') are never chosen as the header.
    """
    non_null = [v for v in row if v is not None and str(v).strip()]
    if not non_null:
        return False
    data_count = sum(1 for v in non_null if _cell_is_data(v))
    return data_count > len(non_null) / 2


def _find_header_row(rows: list[list[Any]], max_scan: int = _MAX_HEADER_SCAN) -> int:
    """Find the best header row in the first *max_scan* rows.

    Scores each row by:
    1. Number of cells that match period patterns (quarterly, monthly, annual)
    2. Number of non-null string values (fallback for non-period headers)

    A row is disqualified if it is predominantly numeric/date data — real banner
    and header rows are text-dominant.  This prevents a leading data row whose
    first cell happens to be an ISO month (e.g. '2025-01') from outscoring the
    actual text header row above it.

    Returns the row index, or -1 if no row has enough non-null values.
    A row needs at least 2 non-null string values to qualify.
    """
    best_idx = -1
    best_period_score = 0
    best_string_score = 0

    for i, row in enumerate(rows[:max_scan]):
        # Skip rows that are predominantly numeric/date — those are data rows,
        # not banner text or column headers.
        if _row_is_predominantly_data(row):
            continue

        period_score = 0
        string_score = 0
        for val in row:
            if val is None:
                continue
            s = str(val)
            if not s.strip():
                continue
            string_score += 1
            if _classify_header(s) is not None:
                period_score += 1

        # Need at least 2 non-null strings to be a header candidate
        if string_score < 2:
            continue

        # Prefer rows with period patterns; break ties by string count
        if (period_score, string_score) > (best_period_score, best_string_score):
            best_period_score = period_score
            best_string_score = string_score
            best_idx = i

    # If no period headers found but row 0 has strings (and is not data), use row 0
    if best_idx == -1 and rows and not _row_is_predominantly_data(rows[0]):
        row0_strings = sum(1 for v in rows[0] if v is not None and str(v).strip())
        if row0_strings >= 2:
            best_idx = 0

    return best_idx


# --- Used-range guard bounds ---
# A formatting-only cell (a style with no value) balloons a sheet's declared
# used-range, and read-only iter_rows dutifully fills the gap with millions of
# empty cells — megabytes of nulls from a sheet with a handful of real rows.
# Below _DEGENERATE_CELL_THRESHOLD the sheet is walked normally and trimmed to
# its populated bounding box unconditionally (lossless — only trailing nulls are
# removed), so a tight sheet is byte-identical and a stray-formatting balloon
# collapses. Above the threshold a bounded walk applies. Bounds sit ~10-100x
# above any realistic founder model (largest observed real extraction is a few
# hundred thousand cells).
_DEGENERATE_CELL_THRESHOLD = 2_000_000  # declared rows*cols that flags a sheet outright
_SPARSE_TRIM_WARN_RATIO = 0.25  # trailing-null trim removing more than this earns a warning
_MAX_SHEET_COLS = 1_024  # columns read on a degenerate sheet
_MAX_SHEET_ROWS = 20_000  # rows walked on a degenerate sheet
_KEPT_CELL_BUDGET = 1_000_000  # rows*cols kept per flagged sheet
_WORKBOOK_KEPT_CELL_BUDGET = 2_000_000  # rows*cols kept across all flagged sheets
_OUTPUT_BYTES_WARN = 25_000_000  # serialized-output size that earns a receipt warning
_EMPTY_ROW_RUN_STOP = 500  # consecutive empty rows that end the scan


def _read_sheet_rows(ws: Any, cell_budget: int) -> tuple[list[list[Any]], list[list[str]], list[str], int]:
    """Read a worksheet's rows/coords, bounding pathological used-ranges.

    Returns (rows, coords, warning_notes, kept_cell_count). Two guards:

    - A sheet whose declared used-range exceeds the absolute cell threshold
      gets a bounded walk (column cap, early stop after a long run of empty
      rows, row cap, cell budget) and is trimmed to its populated bounding
      box. The bounded walk is chosen over re-deriving the true extent
      (reset_dimensions / calculate_dimension) so the sheet is parsed once;
      the trade-off — data beyond the caps is dropped — is disclosed in the
      warning, never silent.
    - A sub-threshold sheet that is still large but almost entirely empty
      (formatting-only bloat) is walked normally and then trimmed to its
      populated bounding box, with a warning.

    Sheets matching neither guard return exactly what the unguarded walk
    produced — byte-identical extraction for ordinary models.
    """
    all_rows: list[list[Any]] = []
    all_coords: list[list[str]] = []  # parallel array of cell coordinates
    declared_rows = ws.max_row or 0
    declared_cols = ws.max_column or 0
    declared_cells = declared_rows * declared_cols

    if declared_cells <= _DEGENERATE_CELL_THRESHOLD:
        last_data_row = -1
        max_data_col = -1
        for row in ws.iter_rows(values_only=False):
            vals = [_safe_value(c.value) for c in row]
            all_rows.append(vals)
            all_coords.append([getattr(c, "coordinate", "") for c in row])
            for j, v in enumerate(vals):
                if v is not None:
                    last_data_row = len(all_rows) - 1
                    max_data_col = max(max_data_col, j)
        # Trim trailing all-null rows/cols to the populated bounding box,
        # unconditionally. Lossless by construction: a populated cell anchors the
        # box, so only trailing nulls are removed — a no-op on tight sheets
        # (byte-identical output) and the fix for stray-formatting bloat that
        # inflates the declared used-range far below the degenerate threshold.
        keep_rows = max(last_data_row + 1, 1)
        keep_cols = max(max_data_col + 1, 1)
        all_rows = [r[:keep_cols] for r in all_rows[:keep_rows]]
        all_coords = [c[:keep_cols] for c in all_coords[:keep_rows]]
        kept_cells = keep_rows * keep_cols
        # Warn only when the trim removed a large fraction of the declared range
        # (stray-formatting bloat); a no-op or minor trim stays silent so
        # ordinary models grow no new warnings.
        notes: list[str] = []
        if declared_cells and (declared_cells - kept_cells) > declared_cells * _SPARSE_TRIM_WARN_RATIO:
            notes.append(
                f"sheet '{ws.title}': declared used-range {declared_rows}x{declared_cols} "
                f"({declared_cells:,} cells) far exceeded the populated region "
                f"({keep_rows}x{keep_cols}); trimmed trailing empty rows/columns"
            )
        return all_rows, all_coords, notes, kept_cells

    last_data_row = -1  # index of the last row carrying any value
    max_data_col = -1  # widest populated column seen
    empty_run = 0
    truncated = False
    scan_stopped = False
    for row in ws.iter_rows(min_col=1, max_col=min(declared_cols, _MAX_SHEET_COLS), values_only=False):
        vals = [_safe_value(c.value) for c in row]
        all_rows.append(vals)
        all_coords.append([getattr(c, "coordinate", "") for c in row])
        row_max = -1
        for j, v in enumerate(vals):
            if v is not None:
                row_max = j
        if row_max >= 0:
            last_data_row = len(all_rows) - 1
            empty_run = 0
            max_data_col = max(max_data_col, row_max)
        else:
            empty_run += 1
            if empty_run >= _EMPTY_ROW_RUN_STOP:
                scan_stopped = True
                break
        if len(all_rows) >= _MAX_SHEET_ROWS:
            truncated = True
            break

    keep_cols = max(max_data_col + 1, 1)
    keep_rows = max(last_data_row + 1, 1)
    if keep_rows * keep_cols > cell_budget:
        keep_rows = max(cell_budget // keep_cols, 1)
        truncated = True
    all_rows = [r[:keep_cols] for r in all_rows[:keep_rows]]
    all_coords = [c[:keep_cols] for c in all_coords[:keep_rows]]

    note = (
        f"sheet '{ws.title}': declared used-range {declared_rows}x{declared_cols} "
        f"(~{declared_cells:,} cells) exceeded the extraction bound; "
        f"kept the populated region ({len(all_rows)}x{keep_cols})"
    )
    if truncated:
        note += "; rows beyond the extraction cap were dropped"
    if scan_stopped:
        note += f"; scan stopped after {_EMPTY_ROW_RUN_STOP} consecutive empty rows"
    if declared_cols > _MAX_SHEET_COLS:
        note += f"; columns beyond {_MAX_SHEET_COLS:,} were not scanned"
    return all_rows, all_coords, [note], len(all_rows) * keep_cols


def _cell_ref_col_key(headers: list[str], j: int) -> str:
    """Collision-proof key for column ``j`` in ``cell_refs.cols``.

    The first occurrence of a header keeps the bare header verbatim
    (unique-header sheets stay byte-identical); a duplicated header's 2nd+
    occurrence gets a ``#N`` suffix. Computable from ``(headers, j)`` alone, so
    the producer here and the consumer in ``validate_extraction._find_cell_ref``
    derive the identical key without the cell coordinate. Without this, two
    columns sharing a header collide (last-duplicate-wins) and a value in an
    earlier duplicate column resolves to a later column's coordinate.

    Twin copy lives in ``validate_extraction.py`` — keep the two in sync.
    """
    h = headers[j]
    occurrence = headers[: j + 1].count(h)
    return h if occurrence == 1 else f"{h}#{occurrence}"


def extract_xlsx(file_path: str) -> dict[str, Any]:
    """Extract data from an Excel file."""
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError:
        print(
            "Error: openpyxl is required for .xlsx files. Install with: pip install openpyxl",
            file=sys.stderr,
        )
        sys.exit(1)

    extraction_warnings: list[str] = []
    # Capture openpyxl load/parse warnings (e.g. unsupported sheet extensions)
    # into the output instead of letting them leak to stderr.
    remaining_budget = _WORKBOOK_KEPT_CELL_BUDGET  # shared across flagged sheets
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        wb = load_workbook(file_path, data_only=True, read_only=True)
        sheets = []
        for ws in wb.worksheets:
            all_rows, all_coords, sheet_notes, kept_cells = _read_sheet_rows(
                ws, min(_KEPT_CELL_BUDGET, max(remaining_budget, 1))
            )
            if sheet_notes:
                extraction_warnings.extend(sheet_notes)
                remaining_budget = max(remaining_budget - kept_cells, 0)

            header_idx = _find_header_row(all_rows)
            if header_idx >= 0:
                raw_header = all_rows[header_idx]
                headers = [str(v) if v is not None else f"col_{j}" for j, v in enumerate(raw_header)]
                rows_data = all_rows[header_idx + 1 :]
                coords_data = all_coords[header_idx + 1 :]
            else:
                # No good header row found — use col_N fallback
                ncols = len(all_rows[0]) if all_rows else 0
                headers = [f"col_{j}" for j in range(ncols)]
                rows_data = all_rows
                coords_data = all_coords

            # Build cell_refs: list of {row_index, label, cols: {col_header: "B5"}}
            # Only build when we have a valid header row
            cell_refs: list[dict[str, Any]] = []
            if header_idx >= 0:
                for row_idx, (row_vals, row_coords) in enumerate(zip(rows_data, coords_data, strict=True)):
                    if not row_vals or row_vals[0] is None:
                        continue
                    row_label = str(row_vals[0]).strip()
                    if not row_label:
                        continue
                    cols: dict[str, str] = {}
                    for j in range(1, min(len(row_vals), len(headers))):
                        if isinstance(row_vals[j], (int, float)) and not isinstance(row_vals[j], bool):
                            cols[_cell_ref_col_key(headers, j)] = row_coords[j] if j < len(row_coords) else ""
                    if cols:
                        cell_refs.append({"row_index": row_idx, "label": row_label, "cols": cols})

            # Rows before the detected header (company name, logos, etc.)
            pre_header: list[list[Any]] = []
            if header_idx > 0:
                pre_header = all_rows[:header_idx]

            sheets.append(
                {
                    "name": ws.title,
                    "headers": headers,
                    "rows": rows_data,
                    "detected_type": _detect_tab_type(ws.title),
                    "periodicity": detect_periodicity(headers),
                    "row_count": len(rows_data),
                    "col_count": len(headers),
                    "pre_header_rows": pre_header,
                    "cell_refs": cell_refs,
                }
            )
        wb.close()

    # De-duplicated openpyxl-origin warnings after the sheet-specific notes;
    # anything else raised in the block is not ours to report as extraction
    # context and is re-emitted to stderr untouched.
    seen: set[str] = set()
    for w in caught:
        msg = str(w.message)
        if "openpyxl" in (getattr(w, "filename", "") or ""):
            if msg not in seen:
                seen.add(msg)
                extraction_warnings.append(msg)
        else:
            print(f"{w.category.__name__}: {msg}", file=sys.stderr)
    result: dict[str, Any] = {
        "sheets": sheets,
        "source_format": "xlsx",
        "source_file": os.path.basename(file_path),
        "periodicity_summary": _periodicity_summary(sheets),
        "structural_errors": _count_structural_errors(sheets),
    }
    if extraction_warnings:
        result["extraction_warnings"] = extraction_warnings
    return result


def extract_csv(file_path: str) -> dict[str, Any]:
    """Extract data from a CSV file."""
    with open(file_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows_raw = list(reader)

    if not rows_raw:
        empty_sheets: list[dict[str, Any]] = [
            {
                "name": "Sheet1",
                "headers": [],
                "rows": [],
                "detected_type": None,
                "periodicity": "unknown",
                "row_count": 0,
                "col_count": 0,
                "cell_refs": [],
            }
        ]
        return {
            "sheets": empty_sheets,
            "source_format": "csv",
            "source_file": os.path.basename(file_path),
            "periodicity_summary": "unknown",
        }

    headers = rows_raw[0]
    rows_data = []
    for row in rows_raw[1:]:
        row_vals: list[Any] = []
        for v in row:
            # Try to coerce to number
            try:
                row_vals.append(int(v))
            except ValueError:
                try:
                    row_vals.append(float(v))
                except ValueError:
                    row_vals.append(v if v else None)
        rows_data.append(row_vals)

    name = os.path.splitext(os.path.basename(file_path))[0]
    csv_sheets = [
        {
            "name": name,
            "headers": headers,
            "rows": rows_data,
            "detected_type": _detect_tab_type(name),
            "periodicity": detect_periodicity(headers),
            "row_count": len(rows_data),
            "col_count": len(headers),
            "pre_header_rows": [],
            "cell_refs": [],
        }
    ]
    return {
        "sheets": csv_sheets,
        "source_format": "csv",
        "source_file": os.path.basename(file_path),
        "periodicity_summary": _periodicity_summary(csv_sheets),
    }


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Extract structured data from financial model files")
    group = p.add_mutually_exclusive_group(required=True)
    group.add_argument("--file", help="Path to .xlsx or .csv file")
    group.add_argument("--stdin", action="store_true", help="Read pre-structured JSON from stdin")
    p.add_argument("--pretty", action="store_true", help="Pretty-print JSON output")
    p.add_argument("-o", "--output", help="Write output to file instead of stdout")
    return p.parse_args()


def main() -> None:
    args = parse_args()

    if args.stdin:
        if sys.stdin.isatty():
            print("Error: --stdin requires piped input", file=sys.stderr)
            sys.exit(1)
        try:
            data = json.load(sys.stdin)
        except json.JSONDecodeError as e:
            print(f"Error: invalid JSON on stdin: {e}", file=sys.stderr)
            sys.exit(1)
    else:
        file_path = args.file
        if not os.path.isfile(file_path):
            print(f"Error: file not found: {file_path}", file=sys.stderr)
            sys.exit(1)

        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".xlsx":
            data = extract_xlsx(file_path)
        elif ext == ".csv":
            data = extract_csv(file_path)
        else:
            print(f"Error: unsupported file type '{ext}' (expected .xlsx or .csv)", file=sys.stderr)
            sys.exit(1)

    # model_data.json is read and grepped downstream (the INPUTS_REVIEW dispatch
    # and the validation cross-reference), so the documented Step-2 invocation
    # passes --pretty to keep it line-navigable (indent=2 ~ one cell region per
    # line). Without --pretty this is a single multi-MB line — never drop
    # --pretty from that invocation.
    indent = 2 if args.pretty else None
    out = json.dumps(data, indent=indent) + "\n"
    summary: dict[str, Any] = {"sheets": len(data.get("sheets", []))}
    warn_list: list[str] = list(data.get("extraction_warnings", []))
    out_bytes = len(out.encode("utf-8"))
    if out_bytes > _OUTPUT_BYTES_WARN:
        # Dense real data is never truncated, but its size is called out.
        size_note = f"extraction output is {out_bytes:,} bytes; review the source before downstream processing"
        warn_list.append(size_note)
        if not args.output:
            print(size_note, file=sys.stderr)
    if len(warn_list) > 10:
        warn_list = warn_list[:10] + [f"... and {len(warn_list) - 10} more (full list in the output file)"]
    if warn_list:
        summary["extraction_warnings"] = warn_list
    _write_output(out, args.output, summary=summary)


if __name__ == "__main__":
    main()
