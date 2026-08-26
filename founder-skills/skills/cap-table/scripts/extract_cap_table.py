#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///
"""Validator + cap-table extraction for Lanes 2/3/4.

Modes:
  * --mode=validate: read existing inputs.json + instruments.json from --dir
    and schema-validate them. Pure check; useful for Lane 4 (founder pastes
    structured JSON) or as a CI gate after extraction.
  * --mode=carta: extract from a Carta XLSX export (Summary Cap Table +
    optional Intermediate / Detailed / per-class Ledgers / Convertible
    Ledger / Equity Incentive Plan sheets). Real Carta column conventions
    verified against a corpus of real exports (see
    references/carta-pulley-mapping.md).
  * --mode=pulley: stub — Pulley column mapping is Phase 1 follow-up
    (no real Pulley exports in the corpus yet to verify against).
  * --mode=freeform-emit: deterministically maps Context-A SPREADSHEET_STRUCTURE_DETECTION
    blocks (stdin) + the --xlsx grid into schema-valid inputs.json + instruments.json.

Carta extractor implementation notes (per real-world corpus):
  * Carta puts a banner in rows 2-3; real headers are in row 5.
  * Per-class ledger naming follows `{ClassPrefix} Ledger`
    (CS Ledger, PS1 Ledger, PA1 Ledger, etc.). Larger exports include
    them all; minimal exports have only Summary / Intermediate / Detailed.
  * Convertible Ledger contains SAFEs (security_id prefix "SAFE*") and
    notes in the same sheet; "Note Block Name" groups by SAFE template.
  * Conversion Discount is stored as percent (e.g. 0.2 = 20%);
    normalized to multiplier form per Gotcha #3.
"""

from __future__ import annotations

import argparse
import io
import json
import math
import os
import re
import sys
import warnings
from typing import Any

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _artifact_writer import load_schema  # noqa: E402
from _cap_table_schema_validator import (  # noqa: E402
    check_misplaced_top_level_keys,
    drop_nulls_on_optional_strings,
    validate,
)


class CartaFingerprintMismatchError(ValueError):
    """Raised by _carta_extract when --mode=carta is explicit but the workbook
    doesn't match Carta's verified sheet fingerprint. Caller (_mode_carta)
    catches this and emits an E_CARTA_FINGERPRINT_MISMATCH blocker receipt
    with exit code 1."""


# Carta export sheet fingerprints (verified against real exports in
# captable corpus). The "primary" check uses just `Summary Cap Table` for
# minimal exports; the "full" check tightens the match.
CARTA_PRIMARY_FINGERPRINT = {"Summary Cap Table"}
CARTA_FULL_FINGERPRINT = {
    "Summary Cap Table",
    "Intermediate Cap Table",
    "Detailed Cap Table",
}
# OCX (Open Cap Table eXcel) — Carta's optional standardized format
CARTA_OCX_FINGERPRINT = {"Capitalization by Stakeholder", "Voting Details", "Context"}
PULLEY_FINGERPRINT = {"Ownership"}
PULLEY_CONTRACT_TABS = {
    "Shares",
    "SAFEs",
    "Convertible Notes",
    "Stock Options",
    "RSAs",
    "RSUs",
    "Warrants",
}

# Carta convertible-ledger column names (row 5 headers; verified)
CARTA_CONVERTIBLE_COLUMNS = [
    "Formatted Security ID",
    "Security ID",
    "Stakeholder Name",
    "Stakeholder Email",
    "Principal",
    "Other Consideration",
    "Interest",
    "Total",
    "Destination",
    "Issue Date",
    "Board Approval Date",
    "Termination Date",
    "Canceled Date",
    "Cancellation Reason",
    "Converted Date",
    "Maturity Date",
    "Interest Rate",
    "Valuation Cap",
    "Conversion Discount",
    "Change In Control Percent",
    "Conversion Trigger",
    "Note Block Name",
]

# Heuristic: a SAFE has security_id prefix matching this pattern; everything
# else in Convertible Ledger is a convertible note.
_SAFE_PREFIX_RE = re.compile(r"^SAFE\d*-\d+$", re.IGNORECASE)

_SCHEMA_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "references",
    "schemas",
)


def _validate_file(path: str, schema_path: str) -> list[str]:
    if not os.path.exists(path):
        return [f"file missing: {path}"]
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    schema = load_schema(schema_path)
    # `null` and omission mean the same thing on an optional bare-string field; normalise before
    # validating so a payload is not rejected for a spelling the schema never signposted.
    drop_nulls_on_optional_strings(data, schema)
    errors = validate(data, schema)
    errors.extend(check_misplaced_top_level_keys(data, os.path.basename(path)))
    return errors


def _mode_validate(args: argparse.Namespace) -> int:
    """Read inputs.json + instruments.json + cap_state.json from --dir and validate."""
    errors_by_file = {}
    for fname, sname in [
        ("inputs.json", "inputs.schema.json"),
        ("instruments.json", "instruments.schema.json"),
        ("cap_state.json", "cap_state.schema.json"),
    ]:
        path = os.path.join(args.dir, fname)
        if os.path.exists(path):
            errs = _validate_file(path, os.path.join(_SCHEMA_DIR, sname))
            if errs:
                errors_by_file[fname] = errs

    if errors_by_file:
        sys.stderr.write("extract_cap_table.py validate: errors found\n")
        for fname, errs in errors_by_file.items():
            sys.stderr.write(f"  {fname}:\n")
            for e in errs:
                sys.stderr.write(f"    - {e}\n")
        print(json.dumps({"ok": False, "errors": errors_by_file}))
        return 1
    print(
        json.dumps(
            {
                "ok": True,
                "validated": list(
                    set(["inputs.json", "instruments.json", "cap_state.json"]) & set(os.listdir(args.dir))
                ),
            }
        )
    )
    return 0


# Legacy binary Excel is an OLE2 Compound File; openpyxl reads OOXML (.xlsx, a zip) only and dies on
# .xls with a cryptic "BadZipFile: File is not a zip file". Detect the OLE2 magic and fail friendly.
_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
_LEGACY_XLS_REMEDY = (
    "This looks like a legacy Excel '.xls' (binary) file, which this tool can't read. Please re-save "
    "it as '.xlsx' in Excel (File → Save As → Excel Workbook .xlsx) and re-upload."
)


class LegacyXlsError(Exception):
    """Raised when a workbook is a legacy OLE2 .xls (binary), not an OOXML .xlsx (zip)."""

    def __init__(self, message: str = _LEGACY_XLS_REMEDY) -> None:
        self.message = message
        super().__init__(message)


def _load_failed_text(e: Exception) -> str:
    """Friendly text for a `load_failed` blocker: a typed error's `.message` (e.g. LegacyXlsError),
    else the generic `Type: detail` form."""
    return getattr(e, "message", None) or f"{type(e).__name__}: {e}"


def _open_xlsx(path: str) -> Any:
    """Open an XLSX via BytesIO bypass (handles .xlsx(N) macOS dupes).

    Reads the bytes once, then peeks the OLE2 magic up front so a legacy `.xls` (even one misnamed
    `.xlsx`) raises a founder-friendly `LegacyXlsError` instead of openpyxl's cryptic BadZipFile."""
    import openpyxl

    with open(path, "rb") as fp:
        raw = fp.read()
    if raw[:8] == _OLE2_MAGIC:
        raise LegacyXlsError()
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        return openpyxl.load_workbook(io.BytesIO(raw), data_only=True)


def _carta_detect(sheet_names: list[str]) -> str | None:
    """Return 'carta' if Carta-default; 'carta_ocx' if OCX; None otherwise."""
    sheets = {s.strip() for s in sheet_names}
    if CARTA_OCX_FINGERPRINT.issubset(sheets):
        return "carta_ocx"
    if CARTA_FULL_FINGERPRINT.issubset(sheets) or CARTA_PRIMARY_FINGERPRINT.issubset(sheets):
        return "carta"
    return None


def _find_header_row(ws: Any, max_scan: int = 8) -> int:
    """Find the row with the most short-string cells (the header row).

    Carta puts a banner in rows 2-3 and headers in row 5. We don't
    hard-code row 5 because some Carta exports may shift; the heuristic
    matches in practice."""
    best_row = 1
    best_score = -1
    for r_idx in range(1, max_scan + 1):
        try:
            row = next(ws.iter_rows(min_row=r_idx, max_row=r_idx, values_only=True), ())
            strs = [str(c).strip() for c in row if c is not None]
            score = sum(1 for s in strs if 2 <= len(s) <= 50 and len(s.split()) <= 6)
            if score > best_score:
                best_score = score
                best_row = r_idx
        except StopIteration:
            break
    return best_row


def _normalize_discount(d: Any) -> tuple[float | None, str | None]:
    """Per Gotcha #3: convert percent to multiplier form."""
    if d is None or d == "":
        return None, None
    try:
        v = float(d)
    except (TypeError, ValueError):
        return None, f"discount value {d!r} not numeric"
    if 0 < v <= 1.0:
        # Could be already-multiplier (0.80 = 20% discount) OR percent-as-fraction (0.20 = 20% discount)
        # Carta stores 0.20 to mean 20% discount (not 80% multiplier). Convert to multiplier form.
        return 1.0 - v, f"Carta discount {v} (= {v * 100:.0f}%) converted to multiplier {1.0 - v:.4f}"
    if 1 < v <= 100:
        return 1.0 - (v / 100.0), f"Carta discount {v}% converted to multiplier {1.0 - v / 100.0:.4f}"
    return None, f"discount value {v} out of expected range"


def _to_iso_date(v: Any) -> str | None:
    """Convert a spreadsheet date cell to ISO 8601 string (or None).

    Module-scope so both the Carta mapper and the Lane-3 freeform mapper share it.
    """
    if v is None or v == "":
        return None
    if hasattr(v, "date"):  # datetime
        result: str = v.date().isoformat()
        return result
    if hasattr(v, "isoformat"):  # date
        result_d: str = v.isoformat()
        return result_d
    return str(v)[:10]


def _infer_safe_form(cap: Any, discount: Any) -> str:
    """Infer the SAFE form from cap + discount presence."""
    has_cap = cap is not None and cap != "" and float(cap) > 0
    has_disc = discount is not None and discount != "" and float(discount) > 0
    if has_cap and has_disc:
        return "cap_plus_discount"
    if has_cap and not has_disc:
        return "yc_postmoney_cap"
    if not has_cap and has_disc:
        return "yc_postmoney_discount"
    return "other"


def _extract_convertible_ledger(ws: Any) -> list[dict[str, Any]]:
    """Parse the Convertible Ledger sheet. Returns a list of raw records."""
    header_row = _find_header_row(ws)
    rows = list(ws.iter_rows(values_only=True))
    if header_row > len(rows):
        return []
    headers = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(rows[header_row - 1])]
    records = []
    for r in rows[header_row:]:
        if not r or all(c is None or c == "" for c in r):
            continue
        rec = dict(zip(headers, r, strict=False))
        records.append(rec)
    return records


def _convertible_record_to_instrument(rec: dict[str, Any], idx: int) -> tuple[str, dict[str, Any], list[str]]:
    """Convert a Carta Convertible Ledger row to instruments.json entry.

    Returns (kind, instrument_dict, warnings) where kind is "safe" or "note".
    """
    warnings_list: list[str] = []
    sec_id = (rec.get("Security ID") or "").strip()
    is_safe = bool(_SAFE_PREFIX_RE.match(sec_id))

    # Skip cancelled / converted records (these are historical)
    converted = rec.get("Converted Date")
    cancelled = rec.get("Canceled Date")
    if converted or cancelled:
        warnings_list.append(f"{sec_id}: skipped (converted={converted!r} cancelled={cancelled!r})")
        return ("skip", {}, warnings_list)

    investor_name = (rec.get("Stakeholder Name") or "").strip()
    issue_date = _to_iso_date(rec.get("Issue Date")) or "1900-01-01"

    principal = rec.get("Principal") or 0
    interest_rate = rec.get("Interest Rate") or 0  # decimal already; e.g. 0.06
    valuation_cap = rec.get("Valuation Cap")
    if valuation_cap == 0:
        valuation_cap = None
    discount_raw = rec.get("Conversion Discount")
    discount_mult, disc_warning = _normalize_discount(discount_raw)
    if disc_warning:
        warnings_list.append(f"{sec_id}: {disc_warning}")
    # SAFE has no maturity; use sentinel. Notes may have a real date.
    maturity_date = _to_iso_date(rec.get("Maturity Date")) or "9999-12-31"

    if is_safe:
        form = _infer_safe_form(valuation_cap, discount_raw)
        # Carta's "Valuation Cap" column does not distinguish pre-money from
        # post-money. We map cap-only to the post-money form (the modern default
        # and the Carta export skew), but a legacy pre-money SAFE would get the
        # wrong company-capitalization denominator (Gotcha #1). Warn and cap
        # confidence at "medium" so a downstream reviewer confirms the vintage.
        if valuation_cap and form == "yc_postmoney_cap":
            warnings_list.append(
                f"{sec_id}: cap mapped to post_money_valuation_cap (Carta export does not "
                f"distinguish pre/post-money). If this is a legacy pre-money SAFE, the "
                f"company-capitalization denominator differs (Gotcha #1) — confirm vintage."
            )
        return (
            "safe",
            {
                "id": f"safe_{idx:03d}",
                "investor_name": investor_name,
                "purchase_amount": float(principal),
                "post_money_valuation_cap": float(valuation_cap) if valuation_cap else None,
                "discount_multiplier": discount_mult,
                "mfn_provision": None,
                "pro_rata_side_letter": None,
                "issuance_date": issue_date,
                "form": form,
                "conversion_price_override": None,
                "source_document": f"carta:{sec_id}",
                "extraction_confidence": "medium",
            },
            warnings_list,
        )
    # Convertible note.
    #
    # The Carta export does not carry a qualified-financing threshold or a
    # maturity-default treatment, and the math is sensitive to both
    # (maturity_default_treatment selects the note_conversion 7-branch path; the
    # QF threshold gates conversion). Do NOT fabricate them — leave null so the
    # math producer surfaces a structural blocker rather than running on a guess.
    # day_count_basis=365 and interest_converts_to_shares=True are standard
    # conventions we keep, but they are still assumptions on a Carta import, so
    # they (and the other unsupplied fields) get receipt warnings. Carta notes
    # are capped at "medium" extraction_confidence for the same reason.
    # Determine interest_rate_type from the Carta row.
    # Carta exports carry a numeric Interest Rate column but no type qualifier.
    # When a rate is present, default to fixed_numeric_simple (simple interest,
    # fixed numeric rate — the most common convention for convertible notes).
    # When no rate, use "none". Both are assumptions; warn so the agent asks.
    assumed_irt = "fixed_numeric_simple" if interest_rate else "none"
    warnings_list.append(
        f"{sec_id}: interest_rate_type assumed {assumed_irt!r} (Carta export carries no rate-type qualifier) "
        f"— confirm with note text"
    )

    for assumed_field, note in (
        ("day_count_basis", "assumed 365 (Carta export carries no day-count basis); confirm with note text"),
        (
            "interest_converts_to_shares",
            "assumed true (Carta export does not state whether accrued interest converts); confirm with note text",
        ),
        (
            "qualified_financing_threshold",
            "left null (Carta export carries no QF threshold) — provide before running conversion math",
        ),
        (
            "maturity_default_treatment",
            "left null (Carta export carries no maturity-default treatment) — provide before running maturity math",
        ),
        (
            "capitalization_denominator",
            "left null (Carta export carries no cap-denominator) — confirm with note text",
        ),
    ):
        warnings_list.append(f"{sec_id}: {assumed_field} {note}")
    return (
        "note",
        {
            "id": f"note_{idx:03d}",
            "investor_name": investor_name,
            "principal": float(principal),
            "annual_interest_rate": float(interest_rate) if interest_rate else None,
            "interest_rate_type": assumed_irt,
            "day_count_basis": 365,
            "compounding_periods_per_year": None,
            "interest_converts_to_shares": True,
            "issuance_date": issue_date,
            "last_interest_event_date": None,
            "valuation_cap": float(valuation_cap) if valuation_cap else None,
            "discount_multiplier": discount_mult,
            "capitalization_denominator": None,
            "capitalization_denominator_policy": "Carta-supplied: confirm with note text",
            "qualified_financing_threshold": None,
            "maturity_date": maturity_date,
            "maturity_default_treatment": None,
            "maturity_conversion_price_override": None,
            "non_qualified_financing_treatment": None,
            "source_document": f"carta:{sec_id}",
            "extraction_confidence": "medium",
        },
        warnings_list,
    )


def _extract_carta_fd_total(rows: list[Any]) -> int | None:
    """Carta's INDEPENDENT printed grand fully-diluted total from the Summary Cap Table rows.

    Locates the 'Fully Diluted Shares' header column (EXACT normalized match, so the sibling
    'Fully Diluted Shares with …' column is not picked), then the grand-total row labeled 'Totals',
    and returns its value in that column. None if either isn't found. Pure (takes row tuples)."""

    def _norm(s: Any) -> str:
        return " ".join(str(s).split()).lower() if s is not None else ""

    fd_col = None
    for row in rows:
        for j, cell in enumerate(row):
            if _norm(cell) == "fully diluted shares":
                fd_col = j
                break
        if fd_col is not None:
            break
    if fd_col is None:
        return None
    for row in rows:
        label = next((c for c in row if c is not None and str(c).strip()), None)
        if label is not None and _norm(label) == "totals":
            val = row[fd_col] if fd_col < len(row) else None
            try:
                return int(round(float(val)))  # type: ignore[arg-type]
            except (TypeError, ValueError):
                return None
    return None


def _carta_extract(xlsx_path: str) -> dict[str, Any]:
    """Extract structured data from a Carta XLSX into our canonical format.

    Returns a dict with `inputs_seed`, `instruments`, `summary_totals`,
    `warnings`, `format`, `sheets_consumed`. Caller persists to JSON.
    """
    wb = _open_xlsx(xlsx_path)
    sheet_names = wb.sheetnames
    format_detected = _carta_detect(sheet_names) or "unknown"
    warnings_list: list[str] = []
    sheets_consumed: list[str] = []

    # When --mode=carta is explicit but the sheet fingerprint doesn't match
    # Carta's verified shape ("Summary Cap Table" + "Convertible Ledger"),
    # fail loudly. The previous behavior returned ok=true, format="unknown"
    # while still running the Convertible Ledger consumer on whatever sheet
    # happened to be named that, producing nonsense (cancelled SAFEs not
    # skipped, discount normalization not applied, both rows classified as
    # notes). Raising E_CARTA_FINGERPRINT_MISMATCH tells the founder the
    # file isn't the expected shape so they can re-export or fall back to
    # Lane 3 (--mode=freeform-emit).
    if format_detected == "unknown":
        raise CartaFingerprintMismatchError(
            "E_CARTA_FINGERPRINT_MISMATCH: --mode=carta was specified but the workbook "
            f"sheet names {sheet_names!r} do not match Carta's verified Summary Cap Table + "
            "Convertible Ledger fingerprint. The export may be from a different version of "
            "Carta, a different vendor (Pulley/etc.), or a custom workbook. Re-export the "
            "Carta cap table, or fall back to Lane 3: dispatch a Context-A "
            "SPREADSHEET_STRUCTURE_DETECTION sub-agent and run --mode=freeform-emit."
        )

    # 1. Read Summary Cap Table for share-class totals
    summary_totals: dict[str, Any] = {}
    fd_total = None
    if "Summary Cap Table" in sheet_names:
        sheets_consumed.append("Summary Cap Table")
        ws = wb["Summary Cap Table"]
        # Extract company name from row 2 (banner: "{Company} Summary Cap Table")
        try:
            row2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
            banner = next((str(c) for c in row2 if c), "")
            company_match = re.match(r"^([^,]+?)\s+Summary Cap Table", banner)
            if company_match:
                summary_totals["company_name"] = company_match.group(1).strip()
        except StopIteration:
            pass
        # Extract as-of date from row 3
        try:
            row3 = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
            row3_banner = next((str(c) for c in row3 if c), "")
            as_of_match = re.match(r"^As of (\S+)", row3_banner)
            if as_of_match:
                date_str = as_of_match.group(1)
                # Carta uses M/D/YYYY format; convert to ISO
                from datetime import datetime

                try:
                    parsed = datetime.strptime(date_str, "%m/%d/%Y")
                    summary_totals["as_of_date"] = parsed.date().isoformat()
                except ValueError:
                    summary_totals["as_of_date"] = date_str
        except StopIteration:
            pass
        # A1: capture Carta's INDEPENDENT printed grand fully-diluted total (the 'Totals' row) so cap_state
        # can cross-foot the computed FD against it. Independent because Carta computes it, not the rebuilt
        # rows — the one non-circular reconciliation anchor.
        fd_total = _extract_carta_fd_total(list(ws.iter_rows(values_only=True)))
        if fd_total is not None:
            summary_totals["fully_diluted"] = fd_total

    # 2. Extract convertibles from Convertible Ledger (if present)
    instruments_safes = []
    instruments_notes = []
    if "Convertible Ledger" in sheet_names:
        sheets_consumed.append("Convertible Ledger")
        ws = wb["Convertible Ledger"]
        records = _extract_convertible_ledger(ws)
        for idx, rec in enumerate(records, start=1):
            kind, instrument, rec_warnings = _convertible_record_to_instrument(rec, idx)
            warnings_list.extend(rec_warnings)
            if kind == "safe":
                instruments_safes.append(instrument)
            elif kind == "note":
                instruments_notes.append(instrument)

    instruments = {
        "safes": instruments_safes,
        "convertible_notes": instruments_notes,
        "warrants": [],  # Lane-2 warrant extraction is a follow-up (per-class warrant ledgers)
        "option_grants": [],  # Lane-2 grant extraction is a follow-up (Equity Incentive Plan)
        "metadata": {},
    }

    return {
        "format": format_detected,
        "sheets_in_workbook": sheet_names,
        "sheets_consumed": sheets_consumed,
        "summary_totals": summary_totals,
        "instruments": instruments,
        "warnings": warnings_list,
        "counts": {
            "safes_extracted": len(instruments_safes),
            "notes_extracted": len(instruments_notes),
            "total_sheets": len(sheet_names),
        },
    }


def _mode_carta(args: argparse.Namespace) -> int:
    """Extract from a Carta XLSX export. Writes to --instruments path
    (appending if file already exists) and emits extraction_audit to -o."""
    if not args.xlsx:
        sys.stderr.write("--xlsx required for --mode=carta\n")
        return 1
    if not os.path.exists(args.xlsx):
        sys.stderr.write(f"file not found: {args.xlsx}\n")
        return 1
    try:
        result = _carta_extract(args.xlsx)
    except CartaFingerprintMismatchError as e:
        # Surface the structured blocker; do NOT run partial consumers.
        err_receipt: dict[str, Any] = {
            "ok": False,
            "mode": "carta",
            "blocker": "E_CARTA_FINGERPRINT_MISMATCH",
            "error": str(e)[:600],
            "remedy": (
                "The workbook isn't the Carta shape we extract from. Re-export from Carta, or "
                "fall back to Lane 3 (--mode=freeform-emit) via a Context-A "
                "SPREADSHEET_STRUCTURE_DETECTION dispatch."
            ),
        }
        print(json.dumps(err_receipt, indent=2))
        return 1
    except LegacyXlsError as e:
        print(json.dumps({"ok": False, "mode": "carta", "blocker": "legacy_xls", "remedy": e.message}, indent=2))
        return 1
    except Exception as e:
        err_receipt = {
            "ok": False,
            "mode": "carta",
            "blocker": "carta_extraction_failed",
            "error": f"{type(e).__name__}: {e}"[:300],
            "remedy": (
                "Fall back to Lane 3 (--mode=freeform-emit) via a Context-A SPREADSHEET_STRUCTURE_DETECTION dispatch."
            ),
        }
        print(json.dumps(err_receipt, indent=2))
        return 1

    receipt: dict[str, Any] = {
        "ok": True,
        "mode": "carta",
        "format": result["format"],
        "summary": result["summary_totals"],
        "counts": result["counts"],
        "warnings": result["warnings"][:20],  # cap stderr output
        "sheets_consumed": result["sheets_consumed"],
    }

    # Optional: write extraction_audit.json with full data
    if args.output:
        audit_path = os.path.abspath(args.output)
        os.makedirs(os.path.dirname(audit_path) or ".", exist_ok=True)
        with open(audit_path, "w", encoding="utf-8") as f:
            json.dump(result, f, indent=2, default=str)
        receipt["audit_written_to"] = audit_path

    # Optional: write instruments.json
    if hasattr(args, "instruments") and args.instruments:
        inst_path = os.path.abspath(args.instruments)
        os.makedirs(os.path.dirname(inst_path) or ".", exist_ok=True)
        # Merge with existing if present
        existing = {}
        if os.path.exists(inst_path):
            with open(inst_path, encoding="utf-8") as f:
                existing = json.load(f)
        existing_meta = dict(existing.get("metadata") or {})
        existing_meta["schema_version"] = "v0.5.0-instruments"
        if getattr(args, "run_id", None):
            existing_meta["run_id"] = args.run_id
        merged = {
            "safes": existing.get("safes", []) + result["instruments"]["safes"],
            "convertible_notes": existing.get("convertible_notes", []) + result["instruments"]["convertible_notes"],
            "warrants": existing.get("warrants", []),
            "option_grants": existing.get("option_grants", []),
            "metadata": existing_meta,
        }
        with open(inst_path, "w", encoding="utf-8") as f:
            json.dump(merged, f, indent=2, default=str)
        receipt["instruments_written_to"] = inst_path

    print(json.dumps(receipt, indent=2))
    if args.pretty and result["warnings"]:
        sys.stderr.write("  warnings:\n")
        for w in result["warnings"][:10]:
            sys.stderr.write(f"    {w}\n")
    return 0


def _mode_pulley_stub(args: argparse.Namespace) -> int:
    """Pulley not yet implemented; route to freeform."""
    receipt = {
        "ok": False,
        "mode": "pulley",
        "blocker": "pulley_mapping_not_yet_implemented",
        "remedy": (
            "Pulley extraction is a Phase 1 follow-up — no real Pulley exports "
            "in the test corpus to verify against. Dispatch the Context-A "
            "SPREADSHEET_STRUCTURE_DETECTION sub-agent and run --mode=freeform-emit (Lane 3)."
        ),
    }
    print(json.dumps(receipt, indent=2))
    return 1


def _serialize_cell(value: Any) -> Any:
    """Convert a cell value to a JSON-serializable form.

    openpyxl data_only=True returns computed values, which may include
    datetime.datetime / datetime.date / datetime.time objects (none of which
    are JSON-serializable) and datetime.timedelta for duration-formatted
    cells. Convert datetimes/times to ISO strings and timedeltas to their
    string form; leave everything else as-is.
    """
    import datetime

    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    if isinstance(value, datetime.timedelta):
        return str(value)
    return value


# --- Lane-3 grid payload compaction (H4: control-frame size cap) -------------
#
# The --mode=grid dump is inlined into the SPREADSHEET_STRUCTURE_DETECTION
# dispatch prompt, which becomes a harness/Cowork control frame with a hard size
# ceiling (256 KiB). A large freeform workbook can blow past that. But the grid
# is consumed ONLY for structure/role detection — the deterministic
# --mode=freeform-emit phase re-reads the FULL grid straight from the file — so
# the structure-detection grid can be trimmed, rounded, and row-elided down to a
# byte budget with no effect on final-output fidelity.

GRID_BUDGET_BYTES = 200_000  # ~195 KiB; headroom under the 256 KiB control-frame cap
_GRID_FLOAT_SIG = 8  # significant figures kept when rounding floats
_GRID_ELIDE_HEAD = 40  # data rows kept at the top of an elided block
_GRID_ELIDE_TAIL = 10  # data rows kept at the bottom of an elided block


def _grid_cell_blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def _used_bounds(rows: list[Any], merged_ranges: list[str]) -> tuple[int, int]:
    """Return (last_row, last_col), 1-based, covering every non-blank cell plus
    any merged range. (0, 0) when the sheet has no content. Used to drop the
    phantom blank padding openpyxl reports beyond the real used range."""
    from openpyxl.utils import range_boundaries  # type: ignore[import-untyped]

    last_row = 0
    last_col = 0
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, cell in enumerate(row, start=1):
            if not _grid_cell_blank(cell):
                last_row = max(last_row, r_idx)
                last_col = max(last_col, c_idx)
    for mr in merged_ranges:
        try:
            _min_col, _min_row, max_col, max_row = range_boundaries(str(mr))
        except Exception:
            continue
        if max_row:
            last_row = max(last_row, int(max_row))
        if max_col:
            last_col = max(last_col, int(max_col))
    return last_row, last_col


def _trim_sheet(raw: dict[str, Any]) -> dict[str, Any]:
    """Trim a raw sheet to its used bounding box (drops phantom trailing rows and
    columns; keeps interior blanks so column index still maps to column letter)."""
    from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

    rows = raw.get("rows", [])
    last_row, last_col = _used_bounds(rows, raw.get("merged_ranges", []))
    trimmed = [list(row[:last_col]) + [None] * max(0, last_col - len(row)) for row in rows[:last_row]]
    out = dict(raw)
    out["rows"] = trimmed
    out["dimensions"] = f"A1:{get_column_letter(last_col)}{last_row}" if last_row and last_col else "A1:A1"
    return out


def _round_sig(x: float, sig: int) -> float:
    if x == 0 or not math.isfinite(x):
        return x
    digits = sig - int(math.floor(math.log10(abs(x)))) - 1
    return round(x, digits)


def _round_floats(rows: list[Any], sig: int = _GRID_FLOAT_SIG) -> list[Any]:
    """Round every float cell to `sig` significant figures; leave ints/strings/None
    untouched. Structure detection never needs 15-digit precision."""
    return [[_round_sig(c, sig) if isinstance(c, float) else c for c in row] for row in rows]


def _elide_sheet(raw: dict[str, Any], head: int = _GRID_ELIDE_HEAD, tail: int = _GRID_ELIDE_TAIL) -> dict[str, Any]:
    """Collapse a tall block: keep `head` rows from the top and `tail` from the
    bottom, replacing the middle with a marker. Kept rows become indexed objects
    ({"r": <1-based row>, "c": [cells]}) so the sub-agent still reports cell_range
    in true spreadsheet coordinates; the marker is {"elided": n, "rows": "a-b"}."""
    rows = raw.get("rows", [])
    n = len(rows)
    if n <= head + tail:
        return raw  # nothing worth eliding
    indexed: list[Any] = [{"r": i + 1, "c": rows[i]} for i in range(head)]
    first_elided, last_elided = head + 1, n - tail
    indexed.append({"elided": last_elided - first_elided + 1, "rows": f"{first_elided}-{last_elided}"})
    indexed.extend({"r": i + 1, "c": rows[i]} for i in range(n - tail, n))
    out = dict(raw)
    out["rows"] = indexed
    out["indexed"] = True
    return out


def _compact_sheets(raw_sheets: dict[str, Any], budget: int) -> tuple[dict[str, Any], dict[str, Any]]:
    """Compact the grid under `budget` bytes via escalating tiers (each applied
    only while still over budget): trim phantom blanks → round floats → elide tall
    sheets (largest first). Returns (sheets, meta) where meta records which tiers
    fired, the final payload size, and whether it is still over budget."""

    def measure(sh: dict[str, Any]) -> int:
        return len(json.dumps({"ok": True, "mode": "grid", "sheets": sh}, separators=(",", ":")))

    applied: list[str] = ["trim"]
    sheets = {name: _trim_sheet(raw) for name, raw in raw_sheets.items()}

    if measure(sheets) > budget:
        applied.append("round_floats")
        for name in sheets:
            sheets[name]["rows"] = _round_floats(sheets[name]["rows"])

    if measure(sheets) > budget:
        applied.append("elide_rows")
        for name in sorted(sheets, key=lambda n: len(json.dumps(sheets[n], default=str)), reverse=True):
            if measure(sheets) <= budget:
                break
            sheets[name] = _elide_sheet(sheets[name])

    payload_bytes = measure(sheets)
    meta = {
        "applied": applied,
        "payload_bytes": payload_bytes,
        "budget_bytes": budget,
        "over_budget": payload_bytes > budget,
    }
    return sheets, meta


def _mode_grid(args: argparse.Namespace) -> int:
    """Dump every sheet of --xlsx as a cell-value grid for Lane-3 dispatch.

    Output shape (to stdout):
      {"ok": true, "mode": "grid",
       "sheets": {
         "<sheet_name>": {
           "dimensions": "<str>",
           "rows": [[...], ...],       // values_only; None for blank cells
           "merged_ranges": ["A4:C4", ...],
           "indexed": true             // present only when rows were elided
         }, ...
       },
       "compaction": {"applied": [...], "payload_bytes": N, "budget_bytes": B, "over_budget": false}}

    The grid is compacted under a byte budget (default GRID_BUDGET_BYTES, override
    with --grid-budget-bytes) so it fits the control-frame cap; an elided sheet's
    rows are indexed objects ({"r","c"}) interleaved with {"elided","rows"} markers.
    If the grid cannot be compacted under budget, a `grid_too_large` blocker is
    returned (exit 1) rather than overflowing the control frame.

    With -o/--output the full JSON is written to the file and a compact
    receipt is emitted to stdout confirming the write path.
    """
    if not args.xlsx:
        sys.stderr.write("--xlsx required for --mode=grid\n")
        return 1
    if not os.path.exists(args.xlsx):
        err: dict[str, Any] = {
            "ok": False,
            "mode": "grid",
            "blocker": "file_not_found",
            "error": f"file not found: {args.xlsx}",
        }
        print(json.dumps(err))
        return 1

    try:
        wb = _open_xlsx(args.xlsx)
    except LegacyXlsError as e:
        print(json.dumps({"ok": False, "mode": "grid", "blocker": "legacy_xls", "remedy": e.message}))
        return 1
    except Exception as e:
        err = {
            "ok": False,
            "mode": "grid",
            "blocker": "load_failed",
            "error": _load_failed_text(e),
        }
        print(json.dumps(err))
        return 1

    raw_sheets: dict[str, Any] = {}
    for ws in wb.worksheets:
        rows = [[_serialize_cell(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
        merged = [str(r) for r in ws.merged_cells.ranges]
        raw_sheets[ws.title] = {
            "dimensions": ws.dimensions,
            "rows": rows,
            "merged_ranges": merged,
        }

    budget = args.grid_budget_bytes if getattr(args, "grid_budget_bytes", None) else GRID_BUDGET_BYTES
    sheets, compaction = _compact_sheets(raw_sheets, budget)

    if compaction["over_budget"]:
        blocker = {
            "ok": False,
            "mode": "grid",
            "blocker": "grid_too_large",
            "compaction": compaction,
            "error": (
                f"freeform grid is {compaction['payload_bytes']} bytes after compaction, over the "
                f"{budget}-byte control-frame budget. Split the workbook into per-sheet files and run "
                "--mode=grid on each, or reconstruct the cap table conversationally (Lane 4)."
            ),
        }
        print(json.dumps(blocker, indent=2 if args.pretty else None))
        return 1

    payload: dict[str, Any] = {
        "ok": True,
        "mode": "grid",
        "sheets": sheets,
        "compaction": compaction,
    }

    if args.output:
        out = os.path.abspath(args.output)
        os.makedirs(os.path.dirname(out) or ".", exist_ok=True)
        with open(out, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2 if args.pretty else None)
        receipt: dict[str, Any] = {
            "ok": True,
            "mode": "grid",
            "written_to": out,
            "sheet_count": len(sheets),
            "compaction": compaction,
        }
        print(json.dumps(receipt, indent=2 if args.pretty else None))
    else:
        print(json.dumps(payload, indent=2 if args.pretty else None))

    return 0


def _normalize_path(path: str) -> tuple[str, str | None]:
    """Return (normalized_path, warning_message).

    Handles real-world artifacts the corpus test surfaced:
      * macOS duplicate-download suffixes: `file.xlsx(1)`, `file.xlsx(2)` →
        treat as `.xlsx`. (openpyxl's extension check rejects them otherwise.)
      * Trailing whitespace in filenames.
    """
    import re

    warning: str | None = None
    norm = path.strip()
    # Strip macOS dupe suffix: ".xlsx(1)" → ".xlsx", ".pdf(2)" → ".pdf", etc.
    m = re.match(r"^(.*\.(?:xlsx|xls|XLSX|XLS|pdf|PDF|docx))\(\d+\)$", norm)
    if m:
        new_norm = m.group(1)
        warning = (
            f"file path normalized: {os.path.basename(path)!r} → "
            f"{os.path.basename(new_norm)!r} (macOS duplicate-download suffix stripped)"
        )
        norm = new_norm
    return norm, warning


def _check_supported_input_type(path: str) -> tuple[bool, str]:
    """Reject obviously-wrong file types with founder-friendly guidance.

    Returns (ok, message). Corpus test surfaced:
      * `.eml` — email files; user forwarded the email instead of detaching
    """
    lower = path.lower()
    if lower.endswith(".eml"):
        return (
            False,
            "This looks like an email file (.eml). Please attach the cap-table "
            "spreadsheet itself (the XLSX/XLS attachment from the email), not "
            "the email containing it.",
        )
    if lower.endswith(".xls"):
        # Legacy binary .xls — caught here (before any mode dispatch) for a friendly remedy; a
        # misnamed .xls is caught later by the OLE2 magic check in _open_xlsx.
        return False, _LEGACY_XLS_REMEDY
    return True, ""


def _mode_freeform_emit(args: argparse.Namespace) -> int:
    """Deterministically map SPREADSHEET_STRUCTURE_DETECTION blocks (stdin) + the
    --xlsx grid into inputs.json + instruments.json under --dir.

    Required-but-unsupplied fields surface as blockers (no fabrication); the founder's
    answers come back via --answer BLOCK.FIELD=VALUE (e.g.
    --answer 0.interest_rate_type=fixed_numeric_simple). On a clean map (no blockers)
    both artifacts are schema-validated and written. Blockers are a GATE (exit 0), not
    an error: the agent resolves them with the founder and re-runs with --answer.
    """
    import freeform_mapper  # lazy: freeform_mapper imports from this module

    if not args.xlsx:
        sys.stderr.write("--xlsx required for --mode=freeform-emit\n")
        return 1
    if not args.dir:
        sys.stderr.write("--dir required for --mode=freeform-emit\n")
        return 1
    if not os.path.exists(args.xlsx):
        print(
            json.dumps(
                {
                    "ok": False,
                    "mode": "freeform-emit",
                    "blocker": "file_not_found",
                    "error": f"file not found: {args.xlsx}",
                }
            )
        )
        return 1

    try:
        payload = json.load(sys.stdin)
    except Exception as e:
        sys.stderr.write(f"freeform-emit: expected {{blocks:[...]}} JSON on stdin ({e})\n")
        return 1
    _is_dict = isinstance(payload, dict)
    blocks = payload.get("blocks") if _is_dict else None
    # Optional: the sheet's own printed grand FD total (non-circular — sourced from the
    # workbook's "Total" cell, not computed by this skill). Passed to map_freeform so
    # inputs.stated_totals is populated for cap_state cross-foot validation.
    stated_total = payload.get("stated_total") if _is_dict else None
    if not isinstance(blocks, list):
        print(
            json.dumps(
                {"ok": False, "mode": "freeform-emit", "blocker": "bad_input", "error": "stdin must be {blocks:[...]}"}
            )
        )
        return 1

    # Build the cell grid from the workbook (same shape as --mode=grid).
    try:
        wb = _open_xlsx(args.xlsx)
    except LegacyXlsError as e:
        print(json.dumps({"ok": False, "mode": "freeform-emit", "blocker": "legacy_xls", "remedy": e.message}))
        return 1
    except Exception as e:
        err_fe = {"ok": False, "mode": "freeform-emit", "blocker": "load_failed", "error": _load_failed_text(e)}
        print(json.dumps(err_fe))
        return 1
    sheets: dict[str, Any] = {}
    for ws in wb.worksheets:
        rows = [[_serialize_cell(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
        sheets[ws.title] = {
            "dimensions": ws.dimensions,
            "rows": rows,
            "merged_ranges": [str(r) for r in ws.merged_cells.ranges],
        }
    grid = {"ok": True, "mode": "grid", "sheets": sheets}

    # Company meta must already exist (Step 2 wrote inputs.json); equity merges into it.
    inputs_path = os.path.join(args.dir, "inputs.json")
    if not os.path.exists(inputs_path):
        print(
            json.dumps(
                {
                    "ok": False,
                    "mode": "freeform-emit",
                    "blocker": "no_inputs_json",
                    "remedy": (
                        "Run Step 2 first to establish company/mode in inputs.json; "
                        "freeform-emit merges equity into it."
                    ),
                }
            )
        )
        return 1
    with open(inputs_path, encoding="utf-8") as f:
        existing_inputs = json.load(f)

    # answerable_blocker_fields is a typed {field: "enum"|"number"|"bool"} map (freeform-role-map.json).
    # A key whose FIELD half isn't in that map is never read by freeform_mapper (silently ignored
    # downstream) — warn loudly here instead, so a founder's answer to a typo'd/renamed field doesn't
    # vanish without a trace.
    _answerable_fields = set(freeform_mapper._load_role_map().get("answerable_blocker_fields", {}).keys())
    answers: dict[str, Any] = {}
    for kv in args.answer or []:
        if "=" not in kv:
            sys.stderr.write(f"--answer must be BLOCK.FIELD=VALUE, got {kv!r}\n")
            return 1
        k, v = kv.split("=", 1)
        k = k.strip()
        field = k.split(".", 1)[1] if "." in k else k
        if field not in _answerable_fields:
            sys.stderr.write(
                f"warning: --answer key {k!r} names field {field!r}, which is not in "
                f"answerable_blocker_fields {sorted(_answerable_fields)} — the mapper will never read this "
                "answer; it will be silently ignored unless the field name is corrected.\n"
            )
        answers[k] = v.strip()

    result = freeform_mapper.map_freeform(
        blocks,
        grid,
        existing_inputs=existing_inputs,
        answers=answers,
        run_id=args.run_id or "",
        stated_total=stated_total,
    )

    if result["blockers"]:
        # A schema/empty blocker (a block carried the wrong field schema — row_range/columns instead of
        # cell_range/column_role_map — or equity blocks mapped 0 records) is NOT founder-answerable: it
        # needs a re-dispatch with the correct field names. Check this FIRST (its field values are
        # disjoint from the off-contract test below, and its reasons never contain "off-contract").
        schema_empty = any(b.get("field") in {"cell_range", "column_role_map", "emit"} for b in result["blockers"])
        # An off-contract blocker (the sub-agent emitted a block_type/role outside the closed
        # vocabulary) is NOT founder-answerable — steer it to a re-dispatch, not an AskUserQuestion.
        off_contract = any(
            b.get("field") == "block_type" or "off-contract" in str(b.get("reason", "")) for b in result["blockers"]
        )
        if schema_empty:
            next_action = (
                "One or more SPREADSHEET_STRUCTURE_DETECTION blocks used the wrong field schema or mapped "
                "zero rows: each block must carry `cell_range` (the DATA rows, e.g. 'A5:F12') and "
                "`column_role_map` (column-letter -> role) — NOT `row_range`/`columns`/`rows`. Re-dispatch "
                "SPREADSHEET_STRUCTURE_DETECTION with the correct field names and ranges that point at the "
                "data rows (not headers/blank rows); do not ask the founder about these."
            )
        elif off_contract:
            next_action = (
                "One or more blocks are off-contract: the SPREADSHEET_STRUCTURE_DETECTION sub-agent used a "
                "block_type or column-role value outside the closed vocabulary (see the contract in "
                "agents/cap-table.md / references/schemas/freeform-role-map.json). Re-dispatch "
                "SPREADSHEET_STRUCTURE_DETECTION and emit ONLY contract block_types/roles — do not ask the "
                "founder about these. Any remaining founder-answerable blockers still use "
                "--answer <block_index>.<field>=<value>."
            )
        else:
            next_action = (
                "Resolve each blocker with the founder via AskUserQuestion, then re-run with "
                "--answer <block_index>.<field>=<value> for the answerable fields."
            )
        print(
            json.dumps(
                {
                    "ok": False,
                    "mode": "freeform-emit",
                    "blockers": result["blockers"],
                    "warnings": result["warnings"],
                    "next_action": next_action,
                },
                indent=2 if args.pretty else None,
            )
        )
        return 0  # a gate, not an error

    errs: dict[str, Any] = {}
    iv = validate(result["inputs"], load_schema(os.path.join(_SCHEMA_DIR, "inputs.schema.json")))
    if iv:
        errs["inputs.json"] = iv
    nv = validate(result["instruments"], load_schema(os.path.join(_SCHEMA_DIR, "instruments.schema.json")))
    nv.extend(check_misplaced_top_level_keys(result["instruments"], "instruments.json"))
    if nv:
        errs["instruments.json"] = nv
    if errs:
        print(json.dumps({"ok": False, "mode": "freeform-emit", "errors": errs}, indent=2 if args.pretty else None))
        return 1

    os.makedirs(args.dir, exist_ok=True)
    for fname, data in (("inputs.json", result["inputs"]), ("instruments.json", result["instruments"])):
        with open(os.path.join(args.dir, fname), "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
    print(
        json.dumps(
            {
                "ok": True,
                "mode": "freeform-emit",
                "written": ["inputs.json", "instruments.json"],
                "dir": os.path.abspath(args.dir),
                "warnings": result["warnings"],
            },
            indent=2 if args.pretty else None,
        )
    )
    return 0


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "--mode",
        required=True,
        choices=["validate", "carta", "pulley", "freeform-emit", "auto", "grid"],
        help=(
            "validate: schema-check existing JSON in --dir; carta: extract from "
            "Carta XLSX (--xlsx); pulley: stub; freeform-emit: map Context-A "
            "SPREADSHEET_STRUCTURE_DETECTION blocks (stdin) into inputs.json/"
            "instruments.json; auto: sniff sheet fingerprint of --xlsx and "
            "dispatch to carta/pulley, or return a remedy pointing at Lane-3 "
            "(freeform-emit) if neither matches; grid: dump all sheets as "
            "a cell-value grid for Lane-3 SPREADSHEET_STRUCTURE_DETECTION dispatch."
        ),
    )
    p.add_argument("--dir", help="Required for --mode=validate")
    p.add_argument("--xlsx", help="Path to XLSX file (for carta/pulley/auto)")
    p.add_argument("--instruments", help="(Carta mode) Where to write/append instruments.json")
    p.add_argument("-o", "--output", help="Where to write extraction_audit.json")
    p.add_argument("--run-id", dest="run_id", help="Run identifier stamped into metadata.run_id")
    p.add_argument(
        "--grid-budget-bytes",
        dest="grid_budget_bytes",
        type=int,
        default=None,
        help=(
            "(grid mode) byte budget for the compacted cell grid; defaults to "
            f"{GRID_BUDGET_BYTES} (headroom under the 256 KiB control-frame cap)."
        ),
    )
    p.add_argument("--pretty", action="store_true")
    p.add_argument(
        "--answer",
        action="append",
        metavar="BLOCK.FIELD=VALUE",
        help="(freeform-emit) founder answer to a blocker, e.g. 0.interest_rate_type=fixed_numeric_simple. Repeatable.",
    )
    args = p.parse_args()

    # Normalize the xlsx path if provided (handle macOS dupe suffixes)
    if args.xlsx:
        normalized, warning = _normalize_path(args.xlsx)
        if warning:
            sys.stderr.write(f"  note: {warning}\n")
        # Reject unsupported types
        ok, msg = _check_supported_input_type(normalized)
        if not ok:
            print(json.dumps({"ok": False, "mode": args.mode, "blocker": "unsupported_input_type", "remedy": msg}))
            return 1
        args.xlsx = normalized

    if args.mode == "validate":
        if not args.dir:
            sys.stderr.write("--dir required for --mode=validate\n")
            return 1
        return _mode_validate(args)
    if args.mode == "auto":
        # Sniff the workbook and dispatch
        if not args.xlsx:
            sys.stderr.write("--xlsx required for --mode=auto\n")
            return 1
        try:
            wb = _open_xlsx(args.xlsx)
        except LegacyXlsError as e:
            print(json.dumps({"ok": False, "mode": "auto", "blocker": "legacy_xls", "remedy": e.message}))
            return 1
        except Exception as e:
            print(json.dumps({"ok": False, "mode": "auto", "blocker": "load_failed", "error": _load_failed_text(e)}))
            return 1
        detected = _carta_detect(wb.sheetnames)
        if detected in {"carta", "carta_ocx"}:
            args.mode = "carta"
            return _mode_carta(args)
        # Pulley check
        sheets = {s.strip() for s in wb.sheetnames}
        if PULLEY_FINGERPRINT.issubset(sheets) and PULLEY_CONTRACT_TABS & sheets:
            args.mode = "pulley"
            return _mode_pulley_stub(args)
        # Default: freeform — caller must dispatch Context-A
        print(
            json.dumps(
                {
                    "ok": False,
                    "mode": "auto",
                    "detected_format": "freeform",
                    "remedy": (
                        "Workbook does not match Carta or Pulley fingerprints. "
                        "Dispatch the Context-A SPREADSHEET_STRUCTURE_DETECTION sub-agent and pipe "
                        "its blocks to --mode=freeform-emit (Lane 3)."
                    ),
                    "sheet_names": wb.sheetnames,
                },
                indent=2,
            )
        )
        return 1
    if args.mode == "carta":
        return _mode_carta(args)
    if args.mode == "pulley":
        return _mode_pulley_stub(args)
    if args.mode == "grid":
        return _mode_grid(args)
    if args.mode == "freeform-emit":
        return _mode_freeform_emit(args)
    sys.stderr.write(f"unknown mode: {args.mode}\n")
    return 1


if __name__ == "__main__":
    sys.exit(main())
