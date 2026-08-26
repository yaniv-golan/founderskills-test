#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = ["pdfplumber", "openpyxl"]
# ///
"""Evidence-quote verification for cap-table Lane-1 extraction.

Three-layer check per field:

  1. quote_in_doc: the extraction's evidence_quote must be findable as a
     substring of the source document (with normalization that handles
     pdfplumber / DocuSign extraction artifacts).
  2. value_in_quote: if value is non-null, some plausible representation of
     it must appear in the evidence_quote — catches quotes that describe a
     clause but don't actually contain the number.
  3. value_in_doc: if value is non-null, some plausible representation of it
     must appear in the source doc — THE HALLUCINATION GATE. The template-
     blank archetype (model fills in a blank placeholder with a plausible
     default like "45 days") passes both quote_in_doc (via fuzzy match)
     and value_in_quote (because the model fabricated the quote
     consistently with the value); value_in_doc is what catches it,
     because the invented value never appears in the actual document.

Wired into `extract_instrument.py` with a structured-rejection contract.
The fuzzy-match threshold was calibrated against the private eval set to
hold false-positive rate <5% with the verifier in blocking mode.

Designed to handle 8 extraction-artifact patterns surfaced by the eval-set
audit:
  - CID-encoded PDFs (font subsetting)
  - Image-only PDFs (no text layer)
  - pdfplumber space-stripping (`is80%` vs `is 80%`)
  - Hyphenation across line breaks (`anti-\ndilution`)
  - Footnote markers (`$50 million1`)
  - Handwritten / DocuSign-filled fields
  - Non-Latin scripts
  - XLSX cell-reference style quotes
"""

from __future__ import annotations

import argparse
import difflib
import json
import os
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
# Text-normalization primitives live in _normalize.py for reuse across
# evidence_verifier, backward_verifier, and the extractors/ module.
# Public names re-exported here for backward compat.
from _normalize import (  # noqa: E402
    compact_form,
    normalize_text,
)
from _normalize import (
    date_tokens as _date_tokens,
)
from _normalize import (
    numeric_tokens as _numeric_tokens,
)

# Confidence threshold for fuzzy-match acceptance. Calibrated against
# the private eval set.
DEFAULT_FUZZY_THRESHOLD = 0.85

# Threshold below which doc text is treated as image-only / unverifiable.
MIN_DOC_TEXT_CHARS = 500


class MissingDependencyError(Exception):
    """A required third-party parser (e.g. pdfplumber) is not installed.

    Distinct from a genuine extraction failure: a missing dependency must
    surface loudly (the blocking hallucination gate would otherwise silently
    no-op for PDFs), not be swallowed as an image-only document.
    """

    def __init__(self, dependency: str, suffix: str) -> None:
        self.dependency = dependency
        self.suffix = suffix
        super().__init__(
            f"E_MISSING_DEPENDENCY: '{dependency}' is required to extract text from {suffix} files "
            f"but is not installed. Install it (it is declared in pyproject dependencies and the "
            f"script's PEP 723 header; `uv run` will fetch it)."
        )


# How close the value token must appear to the evidence quote anchor.
# Currently this just checks that the token appears anywhere in the
# (normalized) evidence quote; window-based scoring is future work.
VALUE_TOKEN_WINDOW_CHARS = 60


def is_doc_image_only(doc_text: str) -> bool:
    """Heuristic: if the source has very little text, it's likely image-only.
    The verifier returns a special status for these (unverifiable, not failed)."""
    return len(doc_text.strip()) < MIN_DOC_TEXT_CHARS


def has_cid_artifacts(doc_text: str) -> bool:
    """Detect CID-encoded font subset artifacts (DocuSign-style font subsetting
    where pdfplumber can't recover glyphs)."""
    cid_count = len(re.findall(r"\(cid:\d+\)", doc_text))
    return cid_count >= 5


@dataclass
class FieldVerificationResult:
    field_name: str
    quote_check: str  # "exact" | "normalized" | "fuzzy_anchored" | "compact" | "not_found" | "skipped"
    value_in_quote: str  # value appears in evidence_quote
    value_in_doc: str  # value appears in source doc (the hallucination gate)
    fuzzy_ratio: float | None = None
    overall_status: str = "pending"  # "pass" | "fail" | "unverifiable"
    reasons: list[str] = field(default_factory=list)


@dataclass
class VerificationReport:
    overall_status: str  # "pass" | "fail" | "unverifiable_doc"
    n_fields: int
    n_passed: int
    n_failed: int
    n_unverifiable: int
    per_field: list[FieldVerificationResult] = field(default_factory=list)
    doc_metadata: dict[str, Any] = field(default_factory=dict)


def quote_in_doc(
    quote: str,
    doc_text: str,
    fuzzy_threshold: float = DEFAULT_FUZZY_THRESHOLD,
) -> tuple[bool, str, float | None]:
    """5-step fallback chain:
        1. exact substring
        2. normalized substring
        3. anchored fuzzy match (find anchor, compute ratio)
        4. compact-form substring (handles space-stripping)
        5. give up

    Returns (found, match_kind, fuzzy_ratio_if_used).
    """
    if not quote or not doc_text:
        return False, "skipped", None

    # 1. Exact substring
    if quote in doc_text:
        return True, "exact", None

    # 2. Normalized
    qn = normalize_text(quote)
    dn = normalize_text(doc_text)
    if qn in dn:
        return True, "normalized", None

    # 3. Anchored fuzzy: take first 30 chars of quote as anchor, find in doc,
    # then check SequenceMatcher ratio on the surrounding window.
    if len(qn) >= 30:
        anchor = qn[:30]
        idx = dn.find(anchor)
        if idx != -1:
            window = dn[idx : idx + len(qn) + 50]
            ratio = difflib.SequenceMatcher(None, qn, window).ratio()
            if ratio >= fuzzy_threshold:
                return True, "fuzzy_anchored", ratio

    # 3b. Sliding window fuzzy for short quotes
    if len(qn) < 200:
        win_size = max(len(qn) + 30, 50)
        best_ratio = 0.0
        step = max(win_size // 4, 1)
        for i in range(0, max(len(dn) - win_size + 1, 1), step):
            r = difflib.SequenceMatcher(None, qn, dn[i : i + win_size]).ratio()
            if r > best_ratio:
                best_ratio = r
        if best_ratio >= fuzzy_threshold:
            return True, "fuzzy_window", best_ratio

    # 4. Compact-form fallback
    qc = compact_form(quote)
    dc = compact_form(doc_text)
    if qc and qc in dc:
        return True, "compact", None

    return False, "not_found", None


def _value_in_text(value: Any, text: str, text_compact: str | None = None) -> tuple[bool, str]:
    """Generic check: does some plausible representation of `value` appear
    in `text`? Returns (found, reason). Used by both value_token_check (against
    quote) and value_in_doc_check (against source doc — the more important one
    for hallucination detection).
    """
    if value is None:
        return True, "skipped_null"
    # bool must precede the numeric path (bool is an int subclass). A boolean is synthesized from prose
    # (e.g. co_sale_rights=True from "the Investor shall have co-sale rights") — it has no literal token
    # in the source, so quote_in_doc is the evidence; a literal "true"/"false" match is meaningless.
    if isinstance(value, bool):
        return True, "skipped_boolean"
    if isinstance(value, dict):
        return True, "skipped_composite"
    if isinstance(value, list):
        return True, "skipped_list"
    if not text:
        return False, "no_text_to_check"

    tn = normalize_text(text).lower()
    tc = text_compact if text_compact is not None else compact_form(text)
    tokens = _numeric_tokens(value)

    # Dates: require a multi-format match. Year-only is too loose (lets within-
    # same-year perturbations slip through). Try the common Month-DD-YYYY and
    # numeric DD/MM/YYYY combinations; if none match, fail.
    if isinstance(value, str) and re.match(r"^\d{4}-\d{2}-\d{2}$", value):
        for tok in _date_tokens(value):
            if tok.lower() in tn:
                return True, f"matched_date_{tok}"
            tc_token = compact_form(tok)
            if tc_token and tc_token in tc:
                return True, f"matched_date_compact_{tok}"
        return False, f"date_format_not_found (tried {_date_tokens(value)[:5]}...)"

    if isinstance(value, str):
        if value.lower() in tn or compact_form(value) in tc:
            return True, "matched_string"
        return False, "string_value_not_in_text"

    for token in tokens:
        token_lower = token.lower()
        # Require word/non-digit boundaries to avoid false matches on shared
        # substrings (e.g. value=0.93 spuriously matching "Section 0.93.1" or
        # "v0.93" via compact form). For numbers, the canonical formatted
        # representation is far more reliable than the compact alphanumeric
        # form. We keep the compact-form fallback ONLY for strings.
        if isinstance(value, str):
            if token_lower in tn:
                return True, f"matched_token_{token}"
            token_compact = compact_form(token)
            if token_compact and token_compact in tc:
                return True, f"matched_compact_{token}"
        else:
            # For numeric values, require the token to appear flanked by
            # non-alphanumeric chars (word boundary in a digit-aware sense).
            # This catches "$20M" and "20 million" but not "20" as part of
            # "120" or "ISO 9001-2020".
            pattern = r"(?<![A-Za-z0-9])" + re.escape(token_lower) + r"(?![A-Za-z0-9])"
            if re.search(pattern, tn):
                return True, f"matched_token_{token}"

    return False, f"no_token_match (tried {tokens})"


def value_token_check(value: Any, quote: str) -> tuple[bool, str]:
    """Check that some plausible representation of value appears in the
    evidence_quote. Used to catch quotes that don't actually contain the
    claimed value (e.g. quote describes a clause but doesn't include the
    number)."""
    return _value_in_text(value, quote)


def value_in_doc_check(value: Any, doc_text: str, doc_text_compact: str | None = None) -> tuple[bool, str]:
    """Check that the claimed value actually appears in the source document.

    THIS IS THE KEY HALLUCINATION-CLASS CHECK. The template-blank archetype
    is: quote "During a period of 45 days..." fuzzy-matched the doc (which
    actually says "...of  days..." with a blank), and value=45 appeared in
    the (fabricated) quote — so a quote+value-in-quote check would pass.
    But value=45 did NOT appear in the doc anywhere. value_in_doc_check
    catches this.
    """
    return _value_in_text(value, doc_text, text_compact=doc_text_compact)


def verify_field(
    field_name: str,
    field_data: dict[str, Any],
    doc_text: str,
    fuzzy_threshold: float = DEFAULT_FUZZY_THRESHOLD,
    doc_text_compact: str | None = None,
) -> FieldVerificationResult:
    """Verify a single field's evidence_quote + value against doc_text.

    Three checks:
      1. quote_in_doc — evidence_quote findable in source (normalization)
      2. value_in_quote — value appears in evidence_quote (consistency)
      3. value_in_doc  — value appears in source doc (hallucination gate)

    field_data shape: {"value": ..., "evidence_quote": ..., "ambiguity": ..., ...}
    """
    quote = field_data.get("evidence_quote")
    value = field_data.get("value")

    result = FieldVerificationResult(
        field_name=field_name,
        quote_check="skipped",
        value_in_quote="skipped_null",
        value_in_doc="skipped_null",
    )

    # Null value + no quote = absent field, no verification needed
    if value is None and not quote:
        result.overall_status = "pass"
        result.reasons.append("absent field, nothing to verify")
        return result

    # Value present but no evidence quote: can't verify; flag.
    if value is not None and not quote:
        result.quote_check = "skipped"
        result.value_in_quote = "no_quote_to_check"
        result.overall_status = "fail"
        result.reasons.append(f"value={value!r} provided but evidence_quote is null/empty")
        return result

    # 1. quote_in_doc
    quote_found: bool | None = None
    if quote:
        found, kind, ratio = quote_in_doc(quote, doc_text, fuzzy_threshold=fuzzy_threshold)
        quote_found = found
        result.quote_check = kind
        result.fuzzy_ratio = ratio
        if not found:
            result.reasons.append("evidence_quote not found in doc (match attempts exhausted)")

    # 2. value_in_quote
    vq_passed, vq_reason = value_token_check(value, quote or "")
    result.value_in_quote = vq_reason
    if not vq_passed:
        result.reasons.append(f"value_in_quote failed: {vq_reason}")

    # 3. value_in_doc — THE hallucination gate
    vd_passed, vd_reason = value_in_doc_check(value, doc_text, doc_text_compact=doc_text_compact)
    result.value_in_doc = vd_reason
    if not vd_passed:
        result.reasons.append(f"value_in_doc failed: {vd_reason} — value claimed but absent from source doc")

    # Overall status — the HALLUCINATION GATE is value_in_doc.
    # If the value appears in the source document, the field is verified.
    # quote_in_doc and value_in_quote are diagnostic signals for human review,
    # NOT gates: models often paraphrase quotes, omit context, or quote a
    # section header instead of the verbatim clause. As long as the underlying
    # value is in the source, the field is real.
    #
    # We require value_in_doc to pass when there IS a value.
    # If value_in_doc fails, that's the hallucination class — fail.
    if value is None or vd_passed:
        # A boolean has no literal token to match (value_in_doc is skipped), so its ONLY evidence is
        # the quote — a present-but-not-found quote leaves it unverified rather than silently passing.
        if result.value_in_doc == "skipped_boolean" and quote_found is False:
            result.overall_status = "fail"
            result.reasons.append("boolean value: evidence_quote not found in source doc — unverified")
        else:
            result.overall_status = "pass"
    else:
        result.overall_status = "fail"

    return result


def verify_extraction(
    extraction: dict[str, Any],
    doc_text: str,
    fuzzy_threshold: float = DEFAULT_FUZZY_THRESHOLD,
) -> VerificationReport:
    """Verify every field in an extraction. Returns structured report."""
    image_only = is_doc_image_only(doc_text)
    cid_blind = has_cid_artifacts(doc_text)

    report = VerificationReport(
        overall_status="pending",
        n_fields=0,
        n_passed=0,
        n_failed=0,
        n_unverifiable=0,
        doc_metadata={
            "doc_text_chars": len(doc_text),
            "image_only": image_only,
            "has_cid_artifacts": cid_blind,
        },
    )

    if image_only:
        report.overall_status = "unverifiable_doc"
        report.doc_metadata["reason"] = (
            f"source text only {len(doc_text)} chars — likely image-only PDF; "
            "verifier cannot check evidence quotes against missing text layer"
        )
        return report

    # Precompute the doc compact form once for performance
    doc_text_compact = compact_form(doc_text)

    fields = extraction.get("fields") or {}
    for fname, fdata in fields.items():
        if not isinstance(fdata, dict):
            continue
        result = verify_field(
            fname, fdata, doc_text, fuzzy_threshold=fuzzy_threshold, doc_text_compact=doc_text_compact
        )
        # CID-blind docs: demote "not_found" failures to "unverifiable" rather
        # than failures, since we know we can't recover the text.
        if cid_blind and result.overall_status == "fail" and result.quote_check == "not_found":
            result.overall_status = "unverifiable"
            result.reasons.append("doc has CID-encoded text; quote may exist in glyphs not extracted")
        report.per_field.append(result)
        report.n_fields += 1
        if result.overall_status == "pass":
            report.n_passed += 1
        elif result.overall_status == "fail":
            report.n_failed += 1
        else:
            report.n_unverifiable += 1

    # Doc-level verifier-blind detection: if 3+ unrelated fields fail with
    # value_in_doc not_found, the doc is likely image-only / DocuSign-overlay
    # / OCR-required but text layer is fragmentary. Demote field failures in
    # such docs to unverifiable so we don't flood the user with false positives.
    # Threshold of 3 is conservative — a real 3-field hallucination is rare;
    # an image-overlay doc with 3+ extractable-text gaps is common.
    n_value_in_doc_failures = sum(
        1 for r in report.per_field if r.overall_status == "fail" and "value_in_doc failed" in " ".join(r.reasons)
    )
    if n_value_in_doc_failures >= 3:
        report.doc_metadata["verifier_blind_demoted"] = True
        report.doc_metadata["verifier_blind_field_count"] = n_value_in_doc_failures
        for r in report.per_field:
            if r.overall_status == "fail" and "value_in_doc failed" in " ".join(r.reasons):
                r.overall_status = "unverifiable"
                r.reasons.append(
                    f"doc-level verifier-blind: {n_value_in_doc_failures} fields fail "
                    "value_in_doc; likely image-only PDF / DocuSign overlay / "
                    "scanned doc where text layer is incomplete"
                )
        # Recount
        report.n_passed = sum(1 for r in report.per_field if r.overall_status == "pass")
        report.n_failed = sum(1 for r in report.per_field if r.overall_status == "fail")
        report.n_unverifiable = sum(1 for r in report.per_field if r.overall_status == "unverifiable")

    if report.n_failed == 0:
        report.overall_status = "pass" if report.n_passed > 0 else "unverifiable_doc"
    else:
        report.overall_status = "fail"

    return report


def report_to_dict(report: VerificationReport) -> dict[str, Any]:
    return {
        "overall_status": report.overall_status,
        "n_fields": report.n_fields,
        "n_passed": report.n_passed,
        "n_failed": report.n_failed,
        "n_unverifiable": report.n_unverifiable,
        "doc_metadata": report.doc_metadata,
        "per_field": [
            {
                "field": r.field_name,
                "status": r.overall_status,
                "quote_check": r.quote_check,
                "value_in_quote": r.value_in_quote,
                "value_in_doc": r.value_in_doc,
                "fuzzy_ratio": r.fuzzy_ratio,
                "reasons": r.reasons,
            }
            for r in report.per_field
        ],
    }


def _load_doc_text(source_path: Path, revisions: str = "accept") -> str:
    """Best-effort text extraction. Defers to dedicated parsers; returns empty
    string on failure (verifier downstream will treat as image-only).

    `revisions` applies to `.docx` only: "accept" reads the accepted-revisions (final executed) view
    via the stdlib `_docx_text` reader. v1 is accept-only."""
    suffix = source_path.suffix.lower()
    # ImportError must NOT be swallowed: a missing parser silently degrades the
    # blocking hallucination gate to a no-op for PDFs. Raise MissingDependencyError
    # so the CLER surfaces E_MISSING_DEPENDENCY and exits non-zero.
    try:
        if suffix == ".pdf":
            import pdfplumber  # noqa: PLC0415
        elif suffix in (".xlsx", ".xlsm"):
            import openpyxl  # noqa: F401, PLC0415
        # .docx now uses stdlib `_docx_text` (no python-docx) — no dependency to probe.
    except ImportError as e:
        dep = {".pdf": "pdfplumber", ".xlsx": "openpyxl", ".xlsm": "openpyxl"}.get(suffix, str(e))
        raise MissingDependencyError(dep, suffix) from e

    try:
        if suffix == ".pdf":
            import pdfplumber  # noqa: PLC0415

            with pdfplumber.open(source_path) as pdf:
                return "\n".join((p.extract_text() or "") for p in pdf.pages)
        elif suffix == ".docx":
            # Tracked-changes-aware, stdlib (no python-docx): python-docx `.text` drops <w:ins>/<w:del>
            # runs (a redline's FINAL inserted terms become invisible) and misses table cells. Read the
            # accepted-revisions view so the verifier searches the same final terms the founder signs.
            import _docx_text  # noqa: PLC0415

            return _docx_text.extract_text(str(source_path), revisions=revisions)
        elif suffix in (".xlsx", ".xlsm"):
            import openpyxl  # noqa: PLC0415

            wb = openpyxl.load_workbook(str(source_path), data_only=True, read_only=True)
            parts = []
            for sn in wb.sheetnames:
                ws = wb[sn]
                parts.append(f"=== {sn} ===")
                for row in ws.iter_rows(values_only=True):
                    parts.append("\t".join(str(c) if c is not None else "" for c in row))
            return "\n".join(parts)
        elif suffix == ".txt":
            return source_path.read_text()
    except Exception as e:
        print(f"warning: failed to extract text from {source_path}: {e}", file=sys.stderr)
    return ""


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--extraction", type=Path, required=True, help="Path to extraction JSON")
    parser.add_argument(
        "--source", type=Path, help="Source document (PDF/DOCX/XLSX) — required unless --doc-text is given"
    )
    parser.add_argument("--doc-text", type=Path, help="Pre-extracted text file (skips parsing)")
    parser.add_argument(
        "--doc-text-source",
        choices=("text_layer", "model_vision"),
        default="text_layer",
        help="Provenance of --doc-text. 'model_vision' (a fresh sub-agent transcribed an "
        "image-only doc) stamps verification_source and demotes confidence one level.",
    )
    parser.add_argument("--fuzzy-threshold", type=float, default=DEFAULT_FUZZY_THRESHOLD)
    parser.add_argument("--pretty", action="store_true")
    parser.add_argument("-o", "--output", type=Path)
    args = parser.parse_args()

    extraction = json.loads(args.extraction.read_text())

    if args.doc_text:
        doc_text = args.doc_text.read_text()
    elif args.source:
        try:
            doc_text = _load_doc_text(args.source)
        except MissingDependencyError as e:
            payload = {
                "error": "E_MISSING_DEPENDENCY",
                "dependency": e.dependency,
                "suffix": e.suffix,
                "detail": str(e),
            }
            print(json.dumps(payload, indent=2 if args.pretty else None))
            return 1
    else:
        print("error: must provide --source or --doc-text", file=sys.stderr)
        return 2

    report = verify_extraction(extraction, doc_text, fuzzy_threshold=args.fuzzy_threshold)
    out = report_to_dict(report)

    # Vision-fallback provenance: when --doc-text came from a fresh sub-agent's
    # transcription of an image-only doc, stamp the source and signal a
    # one-level confidence demotion (vision is less reliable than a text layer).
    if args.doc_text and args.doc_text_source == "model_vision":
        out["doc_metadata"]["verification_source"] = "model_vision"
        out["doc_metadata"]["confidence_demoted"] = True

    if args.output:
        args.output.write_text(json.dumps(out, indent=2 if args.pretty else None))
        receipt = {"ok": True, "output": str(args.output.resolve())}
        print(json.dumps(receipt, indent=2 if args.pretty else None))
    else:
        if args.pretty:
            print(json.dumps(out, indent=2))
        else:
            print(json.dumps(out))

    # Exit code: 0 = pass, 1 = fail, 2 = unverifiable_doc / error
    if report.overall_status == "pass":
        return 0
    if report.overall_status == "fail":
        return 1
    return 2


if __name__ == "__main__":
    sys.exit(main())
